from django.http import HttpResponse
from django.core.exceptions import ImproperlyConfigured
from avocado.conf import OPTIONAL_DEPS
if not OPTIONAL_DEPS['openpyxl']:
    raise ImproperlyConfigured('openpyxl must be installed to use this '
                               'exporter.')

from openpyxl import Workbook
from _base import BaseExporter


class ExcelExporter(BaseExporter):
    short_name = 'Excel'
    long_name = 'Microsoft Excel 2007 Format'

    file_extension = 'xlsx'
    content_type = 'application/vnd.ms-excel'

    preferred_formats = ('excel', 'string')

    def write(self, iterable, buff=None, *args, **kwargs):
        buff = self.get_file_obj(buff)

        # Reference the header
        header = self.header

        wb = Workbook(optimized_write=True)

        ws_data = wb.create_sheet()
        ws_data.title = 'Data'

        # Create the data worksheet
        ws_data.append([f['label'] for f in header])

        for row in iterable:
            ws_data.append(row)

        ws_dict = wb.create_sheet()
        ws_dict.title = 'Data Dictionary'

        # Create the Data Dictionary worksheet
        ws_dict.append((
            'Label',
            'Type',
            'Description',
        ))

        for f in header:
            ws_dict.append((
                f['label'],
                f['type'],
                f['description'],
            ))

        # This hacked up implementation is due to `save_virtual_workbook`
        # not behaving correctly. This function should handle the work
        # https://bitbucket.org/ericgazoni/openpyxl/src/94b05cf9defb9787b4dfbf9e8dca7ba6e0b33d56/openpyxl/writer/excel.py?at=default#cl-154     # noqa
        # however, no data is actually being saved to the worksheets..
        if isinstance(buff, HttpResponse):
            _buff = self.get_file_obj()
            wb.save(_buff)
            buff.content = _buff.getvalue()
            _buff.close()
        else:
            wb.save(buff)

        return buff
