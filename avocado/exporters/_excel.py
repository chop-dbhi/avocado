from openpyxl.workbook import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from _base import BaseExporter

class ExcelExporter(BaseExporter):
    preferred_formats = ('boolean', 'number', 'string')

    def write(self, buff, virtual=True):
        """ Creates an XML based excel spreadsheet
        `buff` - either a file name or a file-like object to be
            written to.
        `virtual` - if true a virtual worksheet will be saved that can
            be passed to a django response. If false, worksheet is saved
            to buff.
        """
        # Create the workbook and sheets
        wb = Workbook(optimized_write=True)
        ws_data = wb.create_sheet(0)
        ws_data.title = 'Data'
        ws_dict = wb.create_sheet(1)
        ws_dict.title = 'Data Dictionary'

        # Create the Data Dictionary Worksheet
        ws_dict.append(('Field Name', 'Data Type', 'Description',
            'Concept Name', 'Concept Discription'))

        for c in self.concepts:
            cfields = c.conceptfields.select_related('field')
            for cfield in cfields:
                field = cfield.field
                ws_dict.append((field.field_name, field.datatype, field.description,
                    c.name, c.description))

        header = []
        # Create the data worksheet
        for i, row_gen in enumerate(self.read()):
            row = []
            for data in row_gen:
                if i == 0:
                    # Build up header row
                    header.extend(data.keys())
                # Add formatted section to the row
                row.extend(data.values())
            # Write headers on first iteration
            if i == 0:
                ws_data.append(header)
            ws_data.append(row)

        # Save the workbook
        if virtual:
            buff = save_virtual_workbook(wb)
        else:
            wb.save(buff)

        return buff
